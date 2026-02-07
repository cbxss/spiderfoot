# -*- coding: utf-8 -*-
# -----------------------------------------------------------------
# Name:         services
# Purpose:      Service layer for SpiderFoot web UI business logic
# -----------------------------------------------------------------
import html
import json
import logging
import multiprocessing as mp
import random
import string
import time
from copy import deepcopy
from io import BytesIO

import openpyxl

from sflib import SpiderFoot
from sfscan import startSpiderFootScanner
from spiderfoot import SpiderFootDb, SpiderFootHelpers

log = logging.getLogger(f"spiderfoot.{__name__}")


def clean_user_input(input_list: list) -> list:
    """Sanitize user input by converting data to HTML entities.

    Args:
        input_list: list of strings to sanitize

    Returns:
        list: sanitized input
    """
    if not isinstance(input_list, list):
        raise TypeError(f"input_list is {type(input_list)}; expected list()")

    ret = []
    for item in input_list:
        if not item:
            ret.append('')
            continue
        c = html.escape(item, True)
        c = c.replace("&amp;", "&").replace("&quot;", "\"")
        ret.append(c)
    return ret


def build_excel(data: list, column_names: list, sheet_name_index: int = 0) -> bytes:
    """Build an Excel workbook from data.

    Args:
        data: rows of data
        column_names: column header names
        sheet_name_index: index of column to use as sheet name

    Returns:
        bytes: Excel workbook content
    """
    row_nums = {}
    workbook = openpyxl.Workbook()
    default_sheet = workbook.active
    column_names.pop(sheet_name_index)
    allowed_sheet_chars = string.ascii_uppercase + string.digits + '_'

    for row in data:
        sheet_name = "".join([c for c in str(row.pop(sheet_name_index)) if c.upper() in allowed_sheet_chars])
        try:
            sheet = workbook[sheet_name]
        except KeyError:
            workbook.create_sheet(sheet_name)
            sheet = workbook[sheet_name]
            for col_num, column_title in enumerate(column_names, 1):
                cell = sheet.cell(row=1, column=col_num)
                cell.value = column_title
            row_nums[sheet_name] = 2

        for col_num, cell_value in enumerate(row, 1):
            cell = sheet.cell(row=row_nums[sheet_name], column=col_num)
            cell.value = cell_value
        row_nums[sheet_name] += 1

    if row_nums:
        workbook.remove(default_sheet)

    workbook._sheets.sort(key=lambda ws: ws.title)

    with BytesIO() as f:
        workbook.save(f)
        f.seek(0)
        return f.read()


class ScanService:
    """Service for scan operations."""

    def __init__(self, config: dict, default_config: dict, logging_queue):
        self.config = config
        self.default_config = default_config
        self.logging_queue = logging_queue

    def list_scans(self) -> list:
        dbh = SpiderFootDb(self.config)
        data = dbh.scanInstanceList()
        retdata = []

        for row in data:
            created = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[3]))
            riskmatrix = {"HIGH": 0, "MEDIUM": 0, "LOW": 0, "INFO": 0}
            correlations = dbh.scanCorrelationSummary(row[0], by="risk")
            if correlations:
                for c in correlations:
                    riskmatrix[c[0]] = c[1]

            if row[4] == 0:
                started = "Not yet"
            else:
                started = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[4]))

            if row[5] == 0:
                finished = "Not yet"
            else:
                finished = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[5]))

            retdata.append([row[0], row[1], row[2], created, started, finished, row[6], row[7], riskmatrix])

        return retdata

    def scan_status(self, scan_id: str) -> list:
        dbh = SpiderFootDb(self.config)
        data = dbh.scanInstanceGet(scan_id)

        if not data:
            return []

        created = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(data[2]))
        started = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(data[3]))
        ended = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(data[4]))
        riskmatrix = {"HIGH": 0, "MEDIUM": 0, "LOW": 0, "INFO": 0}
        correlations = dbh.scanCorrelationSummary(scan_id, by="risk")
        if correlations:
            for c in correlations:
                riskmatrix[c[0]] = c[1]

        return [data[0], data[1], created, started, ended, data[5], riskmatrix]

    def scan_config(self, scan_id: str) -> dict:
        dbh = SpiderFootDb(self.config)
        ret = {}

        meta = dbh.scanInstanceGet(scan_id)
        if not meta:
            return ret

        if meta[3] != 0:
            started = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(meta[3]))
        else:
            started = "Not yet"

        if meta[4] != 0:
            finished = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(meta[4]))
        else:
            finished = "Not yet"

        ret['meta'] = [meta[0], meta[1], meta[2], started, finished, meta[5]]
        ret['config'] = dbh.scanConfigGet(scan_id)
        ret['configdesc'] = {}
        for key in list(ret['config'].keys()):
            if ':' not in key:
                globaloptdescs = self.config.get('__globaloptdescs__')
                if globaloptdescs:
                    ret['configdesc'][key] = globaloptdescs.get(key, f"{key} (legacy)")
            else:
                [mod_name, mod_opt] = key.split(':')
                if mod_name not in list(self.config['__modules__'].keys()):
                    continue
                if mod_opt not in list(self.config['__modules__'][mod_name]['optdescs'].keys()):
                    continue
                ret['configdesc'][key] = self.config['__modules__'][mod_name]['optdescs'][mod_opt]

        return ret

    def scan_summary(self, scan_id: str, by: str) -> list:
        retdata = []
        dbh = SpiderFootDb(self.config)

        try:
            scandata = dbh.scanResultSummary(scan_id, by)
        except Exception:
            return retdata

        try:
            statusdata = dbh.scanInstanceGet(scan_id)
        except Exception:
            return retdata

        for row in scandata:
            if row[0] == "ROOT":
                continue
            lastseen = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[2]))
            retdata.append([row[0], row[1], lastseen, row[3], row[4], statusdata[5]])

        return retdata

    def scan_correlations(self, scan_id: str) -> list:
        retdata = []
        dbh = SpiderFootDb(self.config)

        try:
            corrdata = dbh.scanCorrelationList(scan_id)
        except Exception:
            return retdata

        for row in corrdata:
            retdata.append([row[0], row[1], row[2], row[3], row[4], row[5], row[6], row[7]])

        return retdata

    def scan_event_results(self, scan_id: str, event_type: str = None, filterfp: bool = False, correlation_id: str = None) -> list:
        retdata = []
        dbh = SpiderFootDb(self.config)

        if not event_type:
            event_type = 'ALL'

        try:
            data = dbh.scanResultEvent(scan_id, event_type, filterfp, correlationId=correlation_id)
        except Exception:
            return retdata

        for row in data:
            lastseen = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[0]))
            retdata.append([
                lastseen,
                html.escape(row[1]),
                html.escape(row[2]),
                row[3], row[5], row[6], row[7], row[8],
                row[13], row[14], row[4]
            ])

        return retdata

    def scan_event_results_unique(self, scan_id: str, event_type: str, filterfp: bool = False) -> list:
        dbh = SpiderFootDb(self.config)
        retdata = []

        try:
            data = dbh.scanResultEventUnique(scan_id, event_type, filterfp)
        except Exception:
            return retdata

        for row in data:
            escaped = html.escape(row[0])
            retdata.append([escaped, row[1], row[2]])

        return retdata

    def scan_history(self, scan_id: str) -> list:
        dbh = SpiderFootDb(self.config)
        try:
            return dbh.scanResultHistory(scan_id)
        except Exception:
            return []

    def scan_logs(self, scan_id: str, limit: str = None, row_id: str = None, reverse: str = None) -> list:
        dbh = SpiderFootDb(self.config)
        retdata = []

        try:
            data = dbh.scanLogs(scan_id, limit, row_id, reverse)
        except Exception:
            return retdata

        for row in data:
            generated = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[0] / 1000))
            retdata.append([generated, row[1], row[2], html.escape(row[3]), row[4]])

        return retdata

    def scan_errors(self, scan_id: str, limit: str = None) -> list:
        dbh = SpiderFootDb(self.config)
        retdata = []

        try:
            data = dbh.scanErrors(scan_id, limit)
        except Exception:
            return retdata

        for row in data:
            generated = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[0] / 1000))
            retdata.append([generated, row[1], html.escape(str(row[2]))])

        return retdata

    def scan_element_type_discovery(self, scan_id: str, event_type: str) -> dict:
        dbh = SpiderFootDb(self.config)
        pc = {}
        datamap = {}
        retdata = {}

        try:
            leaf_set = dbh.scanResultEvent(scan_id, event_type)
            [datamap, pc] = dbh.scanElementSourcesAll(scan_id, leaf_set)
        except Exception:
            return retdata

        del pc['ROOT']
        retdata['tree'] = SpiderFootHelpers.dataParentChildToTree(pc)
        retdata['data'] = datamap
        return retdata

    def start_scan(self, scanname: str, scantarget: str, modulelist: str, typelist: str, usecase: str) -> tuple:
        """Start a new scan.

        Returns:
            tuple: (success: bool, message: str, scan_id: str or None)
        """
        scanname = clean_user_input([scanname])[0]
        scantarget = clean_user_input([scantarget])[0]

        if not scanname:
            return (False, "Invalid request: scan name was not specified.", None)

        if not scantarget:
            return (False, "Invalid request: scan target was not specified.", None)

        if not typelist and not modulelist and not usecase:
            return (False, "Invalid request: no modules specified for scan.", None)

        target_type = SpiderFootHelpers.targetTypeFromString(scantarget)
        if target_type is None:
            return (False, "Invalid target type. Could not recognize it as a target SpiderFoot supports.", None)

        dbh = SpiderFootDb(self.config)
        cfg = deepcopy(self.config)
        sf = SpiderFoot(cfg)

        modlist = []

        if modulelist:
            modlist = modulelist.replace('module_', '').split(',')

        if len(modlist) == 0 and typelist:
            typesx = typelist.replace('type_', '').split(',')
            modlist = sf.modulesProducing(typesx)
            newmods = deepcopy(modlist)
            newmodcpy = deepcopy(newmods)

            while len(newmodcpy) > 0:
                for etype in sf.eventsToModules(newmodcpy):
                    xmods = sf.modulesProducing([etype])
                    for mod in xmods:
                        if mod not in modlist:
                            modlist.append(mod)
                            newmods.append(mod)
                newmodcpy = deepcopy(newmods)
                newmods = []

        if len(modlist) == 0 and usecase:
            for mod in self.config['__modules__']:
                if usecase == 'all' or usecase in self.config['__modules__'][mod]['group']:
                    modlist.append(mod)

        if not modlist:
            return (False, "Invalid request: no modules specified for scan.", None)

        if "sfp__stor_db" not in modlist:
            modlist.append("sfp__stor_db")
        modlist.sort()

        if "sfp__stor_stdout" in modlist:
            modlist.remove("sfp__stor_stdout")

        if target_type in ["HUMAN_NAME", "USERNAME", "BITCOIN_ADDRESS"]:
            scantarget = scantarget.replace("\"", "")
        else:
            scantarget = scantarget.lower()

        scan_id = SpiderFootHelpers.genScanInstanceId()
        try:
            p = mp.Process(target=startSpiderFootScanner, args=(self.logging_queue, scanname, scan_id, scantarget, target_type, modlist, cfg))
            p.daemon = True
            p.start()
        except Exception as e:
            log.error(f"[-] Scan [{scan_id}] failed: {e}")
            return (False, f"[-] Scan [{scan_id}] failed: {e}", None)

        while dbh.scanInstanceGet(scan_id) is None:
            log.info("Waiting for the scan to initialize...")
            time.sleep(1)

        return (True, "SUCCESS", scan_id)

    def stop_scan(self, scan_id: str) -> tuple:
        """Stop a scan.

        Returns:
            tuple: (success: bool, error_code: str, message: str)
        """
        if not scan_id:
            return (False, '404', "No scan specified")

        dbh = SpiderFootDb(self.config)
        ids = scan_id.split(',')

        for sid in ids:
            res = dbh.scanInstanceGet(sid)
            if not res:
                return (False, '404', f"Scan {sid} does not exist")

            scan_status = res[5]
            if scan_status == "FINISHED":
                return (False, '400', f"Scan {sid} has already finished.")
            if scan_status == "ABORTED":
                return (False, '400', f"Scan {sid} has already aborted.")
            if scan_status != "RUNNING" and scan_status != "STARTING":
                return (False, '400', f"The running scan is currently in the state '{scan_status}', please try again later or restart SpiderFoot.")

        for sid in ids:
            dbh.scanInstanceSet(sid, status="ABORT-REQUESTED")

        return (True, '', '')

    def delete_scan(self, scan_id: str) -> tuple:
        """Delete scan(s).

        Returns:
            tuple: (success: bool, error_code: str, message: str)
        """
        if not scan_id:
            return (False, '404', "No scan specified")

        dbh = SpiderFootDb(self.config)
        ids = scan_id.split(',')

        for sid in ids:
            res = dbh.scanInstanceGet(sid)
            if not res:
                return (False, '404', f"Scan {sid} does not exist")
            if res[5] in ["RUNNING", "STARTING", "STARTED"]:
                return (False, '400', f"Scan {sid} is {res[5]}. You cannot delete running scans.")

        for sid in ids:
            dbh.scanInstanceDelete(sid)

        return (True, '', '')

    def rerun_scan(self, scan_id: str) -> tuple:
        """Rerun a scan.

        Returns:
            tuple: (success: bool, message: str, new_scan_id: str or None)
        """
        cfg = deepcopy(self.config)
        dbh = SpiderFootDb(cfg)
        info = dbh.scanInstanceGet(scan_id)

        if not info:
            return (False, "Invalid scan ID.", None)

        scanname = info[0]
        scantarget = info[1]

        scanconfig = dbh.scanConfigGet(scan_id)
        if not scanconfig:
            return (False, f"Error loading config from scan: {scan_id}", None)

        modlist = scanconfig['_modulesenabled'].split(',')
        if "sfp__stor_stdout" in modlist:
            modlist.remove("sfp__stor_stdout")

        target_type = SpiderFootHelpers.targetTypeFromString(scantarget)
        if not target_type:
            target_type = SpiderFootHelpers.targetTypeFromString(f'"{scantarget}"')

        if target_type not in ["HUMAN_NAME", "BITCOIN_ADDRESS"]:
            scantarget = scantarget.lower()

        new_scan_id = SpiderFootHelpers.genScanInstanceId()
        try:
            p = mp.Process(target=startSpiderFootScanner, args=(self.logging_queue, scanname, new_scan_id, scantarget, target_type, modlist, cfg))
            p.daemon = True
            p.start()
        except Exception as e:
            log.error(f"[-] Scan [{new_scan_id}] failed: {e}")
            return (False, f"[-] Scan [{new_scan_id}] failed: {e}", None)

        while dbh.scanInstanceGet(new_scan_id) is None:
            log.info("Waiting for the scan to initialize...")
            time.sleep(1)

        return (True, "SUCCESS", new_scan_id)

    def set_false_positive(self, scan_id: str, result_ids: str, fp: str) -> tuple:
        """Set/unset false positive flag.

        Returns:
            tuple: (status: str, message: str) where status is SUCCESS/WARNING/ERROR
        """
        dbh = SpiderFootDb(self.config)

        if fp not in ["0", "1"]:
            return ("ERROR", "No FP flag set or not set correctly.")

        try:
            ids = json.loads(result_ids)
        except Exception:
            return ("ERROR", "No IDs supplied.")

        status = dbh.scanInstanceGet(scan_id)
        if not status:
            return ("ERROR", f"Invalid scan ID: {scan_id}")

        if status[5] not in ["ABORTED", "FINISHED", "ERROR-FAILED"]:
            return ("WARNING", "Scan must be in a finished state when setting False Positives.")

        if fp == "0":
            data = dbh.scanElementSourcesDirect(scan_id, ids)
            for row in data:
                if str(row[14]) == "1":
                    return ("WARNING", f"Cannot unset element {scan_id} as False Positive if a parent element is still False Positive.")

        childs = dbh.scanElementChildrenAll(scan_id, ids)
        all_ids = ids + childs

        ret = dbh.scanResultsUpdateFP(scan_id, all_ids, fp)
        if ret:
            return ("SUCCESS", "")

        return ("ERROR", "Exception encountered.")

    def vacuum_db(self) -> tuple:
        """Vacuum the database.

        Returns:
            tuple: (status: str, message: str)
        """
        dbh = SpiderFootDb(self.config)
        try:
            if dbh.vacuumDB():
                return ("SUCCESS", "")
            return ("ERROR", "Vacuuming the database failed")
        except Exception as e:
            return ("ERROR", f"Vacuuming the database failed: {e}")


class ConfigService:
    """Service for configuration operations."""

    def __init__(self, config: dict, default_config: dict):
        self.config = config
        self.default_config = default_config
        self.token = None

    def get_raw_config(self) -> tuple:
        """Get raw config for the CLI.

        Returns:
            tuple: (token: int, data: dict)
        """
        ret = {}
        self.token = random.SystemRandom().randint(0, 99999999)
        for opt in self.config:
            if not opt.startswith('__'):
                ret["global." + opt] = self.config[opt]
                continue
            if opt == '__modules__':
                for mod in sorted(self.config['__modules__'].keys()):
                    for mo in sorted(self.config['__modules__'][mod]['opts'].keys()):
                        if mo.startswith("_"):
                            continue
                        ret["module." + mod + "." + mo] = self.config['__modules__'][mod]['opts'][mo]

        return (self.token, ret)

    def export_config(self, pattern: str = None) -> str:
        """Export configuration as text."""
        sf = SpiderFoot(self.config)
        conf = sf.configSerialize(self.config)
        content = ""

        for opt in sorted(conf):
            if ":_" in opt or opt.startswith("_"):
                continue
            if pattern:
                if pattern in opt:
                    content += f"{opt}={conf[opt]}\n"
            else:
                content += f"{opt}={conf[opt]}\n"

        return content

    def save_settings(self, allopts: str, token: str, config_file_contents: str = None) -> tuple:
        """Save settings.

        Returns:
            tuple: (success: bool, message: str)
        """
        if str(token) != str(self.token):
            return (False, f"Invalid token ({token})")

        if config_file_contents:
            try:
                tmp = {}
                for line in config_file_contents.split("\n"):
                    if "=" not in line:
                        continue
                    opt_array = line.strip().split("=")
                    if len(opt_array) == 1:
                        opt_array.append("")
                    tmp[opt_array[0]] = '='.join(opt_array[1:])
                allopts = json.dumps(tmp)
            except Exception as e:
                return (False, f"Failed to parse input file. Was it generated from SpiderFoot? ({e})")

        if allopts == "RESET":
            if self.reset_settings():
                return (True, "")
            return (False, "Failed to reset settings")

        try:
            dbh = SpiderFootDb(self.config)
            useropts = json.loads(allopts)
            cleanopts = {}
            for opt in list(useropts.keys()):
                cleanopts[opt] = clean_user_input([useropts[opt]])[0]

            currentopts = deepcopy(self.config)
            sf = SpiderFoot(self.config)
            self.config = sf.configUnserialize(cleanopts, currentopts)
            dbh.configSet(sf.configSerialize(self.config))
        except Exception as e:
            return (False, f"Processing one or more of your inputs failed: {e}")

        return (True, "")

    def save_settings_raw(self, allopts: str, token: str) -> tuple:
        """Save settings (raw/CLI mode).

        Returns:
            tuple: (status: str, message: str)
        """
        if str(token) != str(self.token):
            return ("ERROR", f"Invalid token ({token}).")

        if allopts == "RESET":
            if self.reset_settings():
                return ("SUCCESS", "")
            return ("ERROR", "Failed to reset settings")

        try:
            dbh = SpiderFootDb(self.config)
            useropts = json.loads(allopts)
            cleanopts = {}
            for opt in list(useropts.keys()):
                cleanopts[opt] = clean_user_input([useropts[opt]])[0]

            currentopts = deepcopy(self.config)
            sf = SpiderFoot(self.config)
            self.config = sf.configUnserialize(cleanopts, currentopts)
            dbh.configSet(sf.configSerialize(self.config))
        except Exception as e:
            return ("ERROR", f"Processing one or more of your inputs failed: {e}")

        return ("SUCCESS", "")

    def reset_settings(self) -> bool:
        """Reset settings to default."""
        try:
            dbh = SpiderFootDb(self.config)
            dbh.configClear()
            self.config = deepcopy(self.default_config)
        except Exception:
            return False
        return True


class DataService:
    """Service for data export and search operations."""

    def __init__(self, config: dict):
        self.config = config

    def search(self, scan_id: str = None, event_type: str = None, value: str = None) -> list:
        retdata = []

        if not scan_id and not event_type and not value:
            return retdata

        if not value:
            value = ''

        regex = ""
        if value.startswith("/") and value.endswith("/"):
            regex = value[1:len(value) - 1]
            value = ""

        value = value.replace('*', '%')
        if value in [None, ""] and regex in [None, ""]:
            value = "%"
            regex = ""

        dbh = SpiderFootDb(self.config)
        criteria = {
            'scan_id': scan_id or '',
            'type': event_type or '',
            'value': value or '',
            'regex': regex or '',
        }

        try:
            data = dbh.search(criteria)
        except Exception:
            return retdata

        for row in data:
            lastseen = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[0]))
            escapeddata = html.escape(row[1])
            escapedsrc = html.escape(row[2])
            retdata.append([lastseen, escapeddata, escapedsrc,
                            row[3], row[5], row[6], row[7], row[8], row[10],
                            row[11], row[4], row[13], row[14]])

        return retdata

    def query(self, query: str) -> tuple:
        """Run a SQL query.

        Returns:
            tuple: (success: bool, data_or_error: list|str)
        """
        dbh = SpiderFootDb(self.config)

        if not query:
            return (False, "Invalid query.")

        if not query.lower().startswith("select"):
            return (False, "Non-SELECTs are unpredictable and not recommended.")

        try:
            ret = dbh.dbh.execute(query)
            data = ret.fetchall()
            column_names = [c[0] for c in dbh.dbh.description]
            return (True, [dict(zip(column_names, row)) for row in data])
        except Exception as e:
            return (False, str(e))
